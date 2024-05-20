import sqlite3
from openpyxl import load_workbook

# Buka koneksi ke database
conn = sqlite3.connect('db_mhs.db')
cursor = conn.cursor()

# Baca data dari file Excel
wb = load_workbook('D:\Documents\MPK\Web Publikasi\database siswa\DPIB.xlsx')
sheet = wb.active

# Loop melalui baris Excel dan masukkan data ke database
for row in sheet.iter_rows(min_row=2, values_only=True):
    name, nis, kelas = row
    cursor.execute("INSERT INTO users (name, nis, kelas) VALUES (?, ?, ?)", (name, nis, kelas))

# Commit perubahan dan tutup koneksi
conn.commit()
conn.close()